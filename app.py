import os, re, sqlite3, zipfile, tempfile, json, hashlib
from datetime import datetime, timedelta
from pathlib import Path
from flask import Flask, request, redirect, url_for, render_template, jsonify, flash
import openpyxl
import requests

APP_DIR = Path(__file__).resolve().parent
INSTANCE_DIR = APP_DIR / 'instance'
UPLOAD_DIR = INSTANCE_DIR / 'uploads'
INSTANCE_DIR.mkdir(exist_ok=True)
UPLOAD_DIR.mkdir(exist_ok=True)
DB_PATH = INSTANCE_DIR / 'awp.sqlite3'
CYCLESLOT_PATHS = [APP_DIR / 'cicloslot.json', INSTANCE_DIR / 'cicloslot.json']

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-change-me')
app.config['MAX_CONTENT_LENGTH'] = int(os.environ.get('MAX_UPLOAD_MB', '500')) * 1024 * 1024
DEFAULT_DIVISOR = 100.0


def db():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con


def to_float(v):
    if v is None or v == '':
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def load_cicloslot_json():
    for path in CYCLESLOT_PATHS:
        if path.exists():
            try:
                return json.loads(path.read_text(encoding='utf-8'))
            except Exception:
                return []
    return []


def sync_model_config_from_json():
    data = load_cicloslot_json()
    if not isinstance(data, list):
        return 0
    count = 0
    with db() as con:
        for item in data:
            code = str(item.get('codiceModello') or item.get('codice_modello') or '').strip()
            if not code:
                continue
            name = str(item.get('nomeModello') or item.get('nome_modello') or '').strip()
            ciclo = to_float(item.get('ciclo'))
            payout = to_float(item.get('payout')) or 65
            con.execute('''INSERT OR REPLACE INTO model_config(codice_modello,nome_modello,ciclo,payout,updated_at)
                           VALUES(?,?,?,?,CURRENT_TIMESTAMP)''', (code, name, ciclo, payout))
            count += 1
    return count


def init_db():
    with db() as con:
        con.execute('''
        CREATE TABLE IF NOT EXISTS snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            snapshot_at TEXT NOT NULL,
            source_file TEXT NOT NULL,
            codeid TEXT NOT NULL,
            codeid_provv TEXT,
            stato TEXT,
            data_attivazione TEXT,
            ultimo_collegamento TEXT,
            last_read_at TEXT,
            cnt_in REAL NOT NULL,
            cnt_out REAL NOT NULL,
            sede_code TEXT,
            pdv_aams TEXT,
            sede_nome TEXT,
            indirizzo TEXT,
            comune TEXT,
            provincia TEXT,
            codice_modello TEXT,
            modello TEXT,
            payout_pct REAL,
            em TEXT,
            noe TEXT,
            incasso_giornaliero REAL,
            media_incasso_gg REAL,
            warning TEXT,
            imported_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(snapshot_at, codeid, source_file)
        )''')
        existing = {r['name'] for r in con.execute('PRAGMA table_info(snapshots)')}
        additions = {'data_attivazione':'TEXT','ultimo_collegamento':'TEXT','codice_modello':'TEXT','modello':'TEXT','payout_pct':'REAL','em':'TEXT','noe':'TEXT','incasso_giornaliero':'REAL','media_incasso_gg':'REAL','warning':'TEXT'}
        for col, typ in additions.items():
            if col not in existing:
                con.execute(f'ALTER TABLE snapshots ADD COLUMN {col} {typ}')
        con.execute('''CREATE TABLE IF NOT EXISTS model_config (
            codice_modello TEXT PRIMARY KEY,
            nome_modello TEXT,
            ciclo REAL,
            payout REAL,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )''')
        con.execute('''CREATE TABLE IF NOT EXISTS import_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            imported_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            filename TEXT,
            files_processed INTEGER,
            rows_read INTEGER,
            rows_inserted INTEGER,
            rows_updated INTEGER,
            rows_skipped INTEGER,
            errors TEXT
        )''')
        con.execute('''CREATE TABLE IF NOT EXISTS github_config (
            id INTEGER PRIMARY KEY CHECK(id=1), repo_url TEXT, branch TEXT DEFAULT 'main', folder TEXT DEFAULT '',
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP)''')
        con.execute('''CREATE TABLE IF NOT EXISTS imported_files (
            fingerprint TEXT PRIMARY KEY, filename TEXT NOT NULL, source_type TEXT NOT NULL,
            imported_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP)''')
        con.execute('CREATE INDEX IF NOT EXISTS idx_snap_code_date ON snapshots(codeid, snapshot_at)')
        con.execute('CREATE INDEX IF NOT EXISTS idx_snap_date ON snapshots(snapshot_at)')
        con.execute('CREATE INDEX IF NOT EXISTS idx_snap_location ON snapshots(sede_code, comune)')
        con.execute('CREATE INDEX IF NOT EXISTS idx_snap_model ON snapshots(modello)')


def parse_date_from_filename(name: str):
    base = Path(name).stem
    patterns = [r'(20\d{2})[-_](\d{1,2})[-_](\d{1,2})(?:[-_](\d{1,2})[_-]?(\d{2}))?', r'(\d{1,2})[-_](\d{1,2})[-_](20\d{2})(?:[-_](\d{1,2})[_-]?(\d{2}))?']
    for p in patterns:
        m = re.search(p, base)
        if not m:
            continue
        g = m.groups()
        try:
            if len(g[0]) == 4:
                y, mo, d = int(g[0]), int(g[1]), int(g[2]); hh, mm = int(g[3] or 0), int(g[4] or 0)
            else:
                d, mo, y = int(g[0]), int(g[1]), int(g[2]); hh, mm = int(g[3] or 0), int(g[4] or 0)
            return datetime(y, mo, d, hh, mm)
        except ValueError:
            pass
    return datetime.now()


def norm_header(v):
    if v is None:
        return ''
    s = str(v).replace('\n', ' ').replace('\r', ' ').strip().upper()
    s = re.sub(r'\s+', ' ', s)
    return s


def find_header_row(ws):
    # I sinottici normalmente hanno l'intestazione alla riga 1, ma questa funzione
    # rende il parser robusto anche se in futuro viene aggiunta una riga titolo.
    for row_idx in range(1, min(ws.max_row, 15) + 1):
        values = [norm_header(c.value) for c in ws[row_idx]]
        if 'CODEID' in values and 'CNTTOTIN' in values and 'CNTTOTOT' in values:
            return row_idx, values
    return 1, [norm_header(c.value) for c in ws[1]]


def get_by_alias(row, cols, aliases):
    for name in aliases:
        i = cols.get(norm_header(name))
        if i is not None and i < len(row):
            return row[i]
    return None


def read_xlsx_rows(path: Path, source_name: str, divisor: float):
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    header_row, header_norm = find_header_row(ws)
    cols = {h: i for i, h in enumerate(header_norm) if h}
    missing = [c for c in ['CODEID', 'CNTTOTIN', 'CNTTOTOT'] if c not in cols]
    if missing:
        raise ValueError(f'{source_name}: colonne mancanti {missing}. Colonne lette: {list(cols.keys())[:20]}')
    snapshot_dt = parse_date_from_filename(source_name).isoformat(timespec='minutes')

    def get(row, name):
        i = cols.get(norm_header(name))
        return row[i] if i is not None and i < len(row) else None

    def geta(row, *aliases):
        return get_by_alias(row, cols, aliases)

    parsed = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not get(row, 'CODEID'):
            continue
        parsed.append({
            'snapshot_at': snapshot_dt,
            'source_file': source_name,
            'codeid': str(get(row, 'CODEID')).strip(),
            'codeid_provv': str(get(row, 'CODEID PROVV') or '').strip(),
            'stato': str(get(row, 'DESCR. STATO') or '').strip(),
            'data_attivazione': str(get(row, 'DATA ATTIVAZIONE') or '').strip(),
            'ultimo_collegamento': str(get(row, 'DATA ULTIMO COLLEGAMENTO') or '').strip(),
            'last_read_at': str(get(row, 'DATA ULTIMA LETTURA VAL.') or '').strip(),
            'cnt_in': to_float(get(row, 'CNTTOTIN')) / divisor,
            'cnt_out': to_float(get(row, 'CNTTOTOT')) / divisor,
            'sede_code': str(get(row, 'CODICE SEDE') or '').strip(),
            'pdv_aams': str(get(row, 'CODICE PDV AAMS') or '').strip(),
            'sede_nome': str(get(row, 'DENOMIN. SEDE') or '').strip(),
            'indirizzo': str(get(row, 'INDIRIZZO') or '').strip(),
            'comune': str(get(row, 'COMUNE') or '').strip(),
            'provincia': str(get(row, 'PROVINCIA') or '').strip(),
            'codice_modello': str(geta(row, 'CODICE MODELLO', 'COD. MODELLO', 'COD MODELLO', 'CODICEMODELLO', 'CODICE MOD.') or '').strip(),
            'modello': str(geta(row, 'MODELLO', 'NOME MODELLO', 'DESCRIZIONE MODELLO', 'MODEL') or '').strip(),
            'payout_pct': to_float(get(row, '% OUT')),
            'em': str(get(row, 'E/M') or '').strip(),
            'noe': str(get(row, 'NOE') or '').strip(),
            'incasso_giornaliero': to_float(get(row, 'INCASSO GIORNALIERO')),
            'media_incasso_gg': to_float(get(row, 'MEDIA INCASSO GG ESERCIZIO')),
            'warning': str(get(row, 'WARNING') or '').strip(),
        })
    return parsed


def get_model_config_map(con):
    return {str(r['codice_modello']).strip(): dict(r) for r in con.execute('SELECT * FROM model_config')}


def import_rows(rows):
    inserted = updated = skipped = 0
    fields = ['snapshot_at','source_file','codeid','codeid_provv','stato','data_attivazione','ultimo_collegamento','last_read_at','cnt_in','cnt_out','sede_code','pdv_aams','sede_nome','indirizzo','comune','provincia','codice_modello','modello','payout_pct','em','noe','incasso_giornaliero','media_incasso_gg','warning']
    with db() as con:
        cfg = get_model_config_map(con)
        for r in rows:
            cm = str(r.get('codice_modello') or '').strip()
            if cm in cfg:
                r['modello'] = r.get('modello') or cfg[cm].get('nome_modello')
                if not r.get('payout_pct'):
                    r['payout_pct'] = cfg[cm].get('payout')

            existing_same = con.execute('''
                SELECT id FROM snapshots
                WHERE codeid=? AND COALESCE(last_read_at,'')=COALESCE(?, '')
                  AND ROUND(cnt_in,2)=ROUND(?,2) AND ROUND(cnt_out,2)=ROUND(?,2)
                LIMIT 1
            ''', (r.get('codeid'), r.get('last_read_at'), r.get('cnt_in'), r.get('cnt_out'))).fetchone()
            if existing_same:
                con.execute('''UPDATE snapshots SET
                    codice_modello=COALESCE(NULLIF(?,''), codice_modello),
                    modello=COALESCE(NULLIF(?,''), modello),
                    sede_nome=COALESCE(NULLIF(?,''), sede_nome),
                    comune=COALESCE(NULLIF(?,''), comune),
                    provincia=COALESCE(NULLIF(?,''), provincia),
                    payout_pct=COALESCE(NULLIF(?,0), payout_pct)
                    WHERE id=?
                ''', (r.get('codice_modello'), r.get('modello'), r.get('sede_nome'), r.get('comune'), r.get('provincia'), r.get('payout_pct'), existing_same['id']))
                skipped += 1
                continue

            placeholders = ','.join(['?']*len(fields))
            updates = ','.join([f'{k}=excluded.{k}' for k in fields if k not in ('snapshot_at','codeid','source_file')])
            cur = con.execute(f'''
                INSERT INTO snapshots ({','.join(fields)}) VALUES ({placeholders})
                ON CONFLICT(snapshot_at, codeid, source_file) DO UPDATE SET {updates}
            ''', tuple(r.get(k) for k in fields))
            if cur.rowcount:
                inserted += 1
            else:
                updated += 1
    return {'inserted': inserted, 'updated': updated, 'skipped': skipped}

def is_sinottico_filename(name: str):
    base = Path(name).name
    return base.lower().startswith("sinottico") and base.lower().endswith((".xlsx", ".zip"))


def file_fingerprint(path: Path):
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def parse_github_repo(repo_url: str):
    value = (repo_url or "").strip().rstrip("/")
    m = re.search(r"github\.com/([^/]+)/([^/]+?)(?:\.git)?$", value)
    if not m:
        raise ValueError("URL repository GitHub non valido. Usa https://github.com/proprietario/repository")
    return m.group(1), m.group(2)


def import_from_github(repo_url: str, branch: str, folder: str, divisor: float):
    owner, repo = parse_github_repo(repo_url)
    branch = (branch or "main").strip()
    folder = (folder or "").strip().strip("/")
    token = os.environ.get("GITHUB_TOKEN", "").strip()
    headers = {"Accept": "application/vnd.github+json", "User-Agent": "AWP-Analytics"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    tree_url = f"https://api.github.com/repos/{owner}/{repo}/git/trees/{branch}?recursive=1"
    resp = requests.get(tree_url, headers=headers, timeout=45)
    if resp.status_code >= 400:
        try: detail = resp.json().get("message", resp.text[:200])
        except Exception: detail = resp.text[:200]
        raise ValueError(f"GitHub ha risposto {resp.status_code}: {detail}")
    candidates=[]
    for item in resp.json().get("tree", []):
        remote=item.get("path", "")
        if item.get("type") != "blob": continue
        if folder and not remote.startswith(folder + "/"): continue
        if Path(remote).name.lower().startswith("sinottico") and remote.lower().endswith((".xlsx", ".zip")):
            candidates.append(remote)
    total={"files_found":len(candidates),"files_downloaded":0,"files_duplicate":0,"files":0,"rows":0,"inserted":0,"updated":0,"skipped":0,"errors":[]}
    with tempfile.TemporaryDirectory() as td:
        for remote in sorted(candidates):
            raw=f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{remote}"
            r=requests.get(raw, headers=headers, timeout=90)
            if r.status_code >= 400:
                total["errors"].append(f"{remote}: download HTTP {r.status_code}")
                continue
            local=Path(td)/Path(remote).name
            local.write_bytes(r.content)
            fp=file_fingerprint(local)
            with db() as con:
                exists=con.execute("SELECT 1 FROM imported_files WHERE fingerprint=?",(fp,)).fetchone()
            if exists:
                total["files_duplicate"] += 1
                continue
            try:
                res=import_file(local, local.name, divisor, source_type="github", register_file=False)
                with db() as con:
                    con.execute("INSERT OR IGNORE INTO imported_files(fingerprint,filename,source_type) VALUES(?,?,?)",(fp,remote,"github"))
                total["files_downloaded"] += 1
                for key in ("files","rows","inserted","updated","skipped"): total[key]+=res.get(key,0)
                total["errors"].extend(res.get("errors",[]))
            except Exception as e:
                total["errors"].append(f"{remote}: {e}")
    with db() as con:
        con.execute("""INSERT INTO github_config(id,repo_url,branch,folder,updated_at) VALUES(1,?,?,?,CURRENT_TIMESTAMP)
        ON CONFLICT(id) DO UPDATE SET repo_url=excluded.repo_url,branch=excluded.branch,folder=excluded.folder,updated_at=CURRENT_TIMESTAMP""",(repo_url,branch,folder))
    return total


def import_file(path: Path, original_name: str, divisor: float, source_type="upload", register_file=True):
    if not is_sinottico_filename(original_name):
        raise ValueError("Il file deve iniziare con Sinottico ed essere .xlsx oppure .zip")
    total_files = total_rows = 0
    inserted = updated = skipped = 0
    errors = []
    fp = file_fingerprint(path)
    if register_file:
        with db() as con:
            if con.execute("SELECT 1 FROM imported_files WHERE fingerprint=?", (fp,)).fetchone():
                return {"files":0,"rows":0,"inserted":0,"updated":0,"skipped":0,"errors":[],"file_duplicate":1}

    def process_xlsx(xlsx_path, display_name):
        nonlocal total_files, total_rows, inserted, updated, skipped
        total_files += 1
        rows = read_xlsx_rows(xlsx_path, display_name, divisor)
        total_rows += len(rows)
        res = import_rows(rows)
        inserted += res['inserted']
        updated += res['updated']
        skipped += res['skipped']

    if original_name.lower().endswith('.zip'):
        with tempfile.TemporaryDirectory() as td:
            with zipfile.ZipFile(path) as z:
                z.extractall(td)
            for xlsx in sorted(Path(td).rglob('*.xlsx')):
                if not xlsx.name.lower().startswith('sinottico'):
                    continue
                try:
                    process_xlsx(xlsx, xlsx.name)
                except Exception as e:
                    errors.append(str(e))
    elif original_name.lower().endswith('.xlsx'):
        process_xlsx(path, original_name)
    else:
        raise ValueError('Carica solo file .xlsx o .zip')

    with db() as con:
        if register_file:
            con.execute("INSERT OR IGNORE INTO imported_files(fingerprint,filename,source_type) VALUES(?,?,?)", (fp, original_name, source_type))
        con.execute('''INSERT INTO import_logs(filename,files_processed,rows_read,rows_inserted,rows_updated,rows_skipped,errors)
                       VALUES(?,?,?,?,?,?,?)''', (original_name,total_files,total_rows,inserted,updated,skipped,' | '.join(errors[:20])))
    return {'files': total_files, 'rows': total_rows, 'inserted': inserted, 'updated': updated, 'skipped': skipped, 'errors': errors, 'file_duplicate': 0}

def summary_totals(con):
    return con.execute('''SELECT COUNT(*) snapshots, COUNT(DISTINCT codeid) machines, COUNT(DISTINCT sede_code) locations, COUNT(DISTINCT modello) models, MIN(snapshot_at) first_date, MAX(snapshot_at) last_date FROM snapshots''').fetchone()


def parse_read_datetime(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ('%d/%m/%Y %H:%M:%S','%d/%m/%Y %H:%M','%Y-%m-%d %H:%M:%S','%Y-%m-%d %H:%M','%d-%m-%Y %H:%M:%S','%d-%m-%Y %H:%M','%Y-%m-%dT%H:%M','%d/%m/%Y'):
        try:
            if '%S' in fmt:
                return datetime.strptime(s[:19], fmt)
            if '%H' in fmt:
                return datetime.strptime(s[:16], fmt)
            return datetime.strptime(s[:10], fmt)
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s.replace('Z',''))
    except Exception:
        return None


def read_status(last_read_at):
    dt = parse_read_datetime(last_read_at)
    now = datetime.now()
    if not dt:
        return {'key':'update','label':'Da aggiornare','detail':'Data lettura valida assente o non leggibile','class':'red'}
    if dt.date() == now.date():
        hours = (now - dt).total_seconds()/3600
        if hours <= 3 and hours >= -1:
            return {'key':'today_3h','label':'Oggi < 3h','detail':dt.strftime('%d/%m/%Y %H:%M'),'class':'green'}
        return {'key':'today_old','label':'Oggi > 3h','detail':dt.strftime('%d/%m/%Y %H:%M'),'class':'yellow'}
    if dt.date() == (now.date() - timedelta(days=1)):
        return {'key':'yesterday','label':'Ieri','detail':dt.strftime('%d/%m/%Y %H:%M'),'class':'orange'}
    return {'key':'update','label':'Da aggiornare','detail':dt.strftime('%d/%m/%Y %H:%M'),'class':'red'}


def sort_sql(sort):
    mapping = {
        'pay_first': 'out_to_realign_now DESC',
        'need_in': 'in_to_realign_no_out DESC',
        'low_counters': 'cnt_in ASC',
        'high_counters': 'cnt_in DESC',
        'location': 'sede_nome ASC, comune ASC',
        'city': 'comune ASC, sede_nome ASC',
        'played_most': 'period_in DESC',
        'margin': 'margine_periodo DESC',
        'cycle_end': 'cycle_remaining ASC'
    }
    return mapping.get(sort, mapping['pay_first'])


@app.route('/')
def index():
    q = request.args.get('q', '').strip()
    sort = request.args.get('sort', 'pay_first')
    params = {}
    filters = []
    if q:
        filters.append('(l.codeid LIKE :q OR l.sede_nome LIKE :q OR l.comune LIKE :q OR COALESCE(mc.nome_modello,l.modello) LIKE :q OR l.codice_modello LIKE :q)')
        params['q'] = f'%{q}%'
    where_extra = ' AND ' + ' AND '.join(filters) if filters else ''
    with db() as con:
        totals = summary_totals(con)
        machines = con.execute(f'''
        WITH latest AS (
          SELECT s.*, ROW_NUMBER() OVER (PARTITION BY codeid ORDER BY snapshot_at DESC, id DESC) rn FROM snapshots s
        ), firstlast AS (
          SELECT codeid, MIN(snapshot_at) first_at, MAX(snapshot_at) last_at FROM snapshots GROUP BY codeid
        ), deltas AS (
          SELECT a.codeid, MAX(0, b.cnt_in-a.cnt_in) period_in, MAX(0, b.cnt_out-a.cnt_out) period_out
          FROM firstlast f JOIN snapshots a ON a.codeid=f.codeid AND a.snapshot_at=f.first_at JOIN snapshots b ON b.codeid=f.codeid AND b.snapshot_at=f.last_at
        )
        SELECT l.codeid, l.sede_code, l.sede_nome, l.comune, l.codice_modello, l.last_read_at, l.stato,
               COALESCE(NULLIF(l.modello,''), mc.nome_modello, 'MODELLO NON CENSITO') modello,
               l.cnt_in, l.cnt_out, COALESCE(mc.ciclo,0) ciclo, COALESCE(mc.payout,l.payout_pct,65) payout_pct,
               ROUND(l.cnt_out/NULLIF(l.cnt_in,0)*100,2) payout_real,
               ROUND((l.cnt_in*(COALESCE(mc.payout,l.payout_pct,65)/100.0))-l.cnt_out,2) out_to_realign_now,
               ROUND(CASE WHEN l.cnt_out>(l.cnt_in*(COALESCE(mc.payout,l.payout_pct,65)/100.0)) THEN (l.cnt_out/(COALESCE(mc.payout,l.payout_pct,65)/100.0))-l.cnt_in ELSE 0 END,2) in_to_realign_no_out,
               ROUND(CASE WHEN COALESCE(mc.ciclo,0)>0 THEN (l.cnt_in % mc.ciclo) ELSE NULL END,2) cycle_pos,
               ROUND(CASE WHEN COALESCE(mc.ciclo,0)>0 THEN mc.ciclo-(l.cnt_in % mc.ciclo) ELSE NULL END,2) cycle_remaining,
               ROUND(d.period_in,2) period_in, ROUND(d.period_out,2) period_out, ROUND(d.period_in-d.period_out,2) margine_periodo
        FROM latest l LEFT JOIN model_config mc ON mc.codice_modello=l.codice_modello LEFT JOIN deltas d ON d.codeid=l.codeid
        WHERE l.rn=1 {where_extra}
        ORDER BY {sort_sql(sort)} LIMIT 700
        ''', params).fetchall()
        machines = [dict(m) for m in machines]
        status_counts = {'today_3h':0, 'today_old':0, 'yesterday':0, 'update':0}
        for m in machines:
            st = read_status(m.get('last_read_at'))
            m['read_status'] = st
            status_counts[st['key']] += 1
        last_import = con.execute('SELECT * FROM import_logs ORDER BY id DESC LIMIT 1').fetchone()
        github_cfg = con.execute('SELECT * FROM github_config WHERE id=1').fetchone()
    return render_template('index.html', totals=totals, machines=machines, q=q, sort=sort, status_counts=status_counts, last_import=last_import, github_cfg=github_cfg)


@app.route('/upload', methods=['POST'])
def upload():
    wants_json = request.headers.get('X-Requested-With') == 'fetch' or 'application/json' in request.headers.get('Accept','')
    f = request.files.get('file')
    divisor = float(request.form.get('divisor') or DEFAULT_DIVISOR)
    if not f or not f.filename:
        if wants_json:
            return jsonify({'ok': False, 'message': 'Nessun file caricato'}), 400
        flash('Nessun file caricato')
        return redirect(url_for('index'))
    if not is_sinottico_filename(f.filename):
        msg = 'Sono accettati soltanto file che iniziano con Sinottico e terminano in .xlsx o .zip'
        if wants_json: return jsonify({'ok': False, 'message': msg}), 400
        flash(msg); return redirect(url_for('index'))
    safe = re.sub(r'[^A-Za-z0-9_.-]+', '_', f.filename)
    dest = UPLOAD_DIR / f'{datetime.now().strftime("%Y%m%d%H%M%S")}_{safe}'
    f.save(dest)
    try:
        res = import_file(dest, f.filename, divisor)
        msg = ('File già importato: nessun dato duplicato è stato aggiunto.' if res.get('file_duplicate') else f"Import completato: {res['files']} sinottici elaborati, {res['rows']} righe lette, {res['inserted']} nuove letture salvate, {res['skipped']} scartate perché duplicate.")
        if res['updated']:
            msg += f" {res['updated']} righe aggiornate."
        if res['errors']:
            msg += ' Errori: ' + ' | '.join(res['errors'][:5])
        if wants_json:
            return jsonify({'ok': True, 'message': msg, **res})
        flash(msg)
    except Exception as e:
        if wants_json:
            return jsonify({'ok': False, 'message': f'Errore import: {e}'}), 500
        flash(f'Errore import: {e}')
    return redirect(url_for('index'))


@app.route("/github-import", methods=["POST"])
def github_import():
    wants_json = request.headers.get("X-Requested-With") == "fetch" or "application/json" in request.headers.get("Accept","")
    try:
        res=import_from_github(request.form.get("repo_url",""),request.form.get("branch","main"),request.form.get("folder",""),float(request.form.get("divisor") or DEFAULT_DIVISOR))
        msg=(f"Import GitHub completato: {res['files_found']} file Sinottico trovati, {res['files_downloaded']} nuovi importati, "
             f"{res['files_duplicate']} già presenti, {res['inserted']} nuove letture e {res['skipped']} letture duplicate scartate.")
        if res["errors"]: msg += " Errori: " + " | ".join(res["errors"][:5])
        if wants_json: return jsonify({"ok":True,"message":msg,**res})
        flash(msg)
    except Exception as e:
        if wants_json: return jsonify({"ok":False,"message":str(e)}),500
        flash(f"Errore import GitHub: {e}")
    return redirect(url_for("index"))


@app.route("/analysis")
def analysis():
    filters = {
        "model": (request.args.get("model") or "").strip(),
        "location": (request.args.get("location") or "").strip(),
        "municipality": (request.args.get("municipality") or "").strip(),
        "band": (request.args.get("band") or "").strip().lower(),
        "q": (request.args.get("q") or "").strip(),
    }
    with db() as con:
        rows = con.execute("""
        WITH ordered AS (
          SELECT s.*,
                 ROW_NUMBER() OVER (PARTITION BY codeid ORDER BY snapshot_at DESC,id DESC) rn_desc,
                 LAG(cnt_in) OVER (PARTITION BY codeid ORDER BY snapshot_at,id) prev_in,
                 LAG(cnt_out) OVER (PARTITION BY codeid ORDER BY snapshot_at,id) prev_out
          FROM snapshots s
        ), deltas AS (
          SELECT *,
                 CASE WHEN prev_in IS NULL OR cnt_in < prev_in THEN NULL ELSE cnt_in-prev_in END din,
                 CASE WHEN prev_out IS NULL OR cnt_out < prev_out THEN NULL ELSE cnt_out-prev_out END dout
          FROM ordered
        ), hist AS (
          SELECT codeid,
                 COUNT(din) observations,
                 AVG(CASE WHEN din IS NOT NULL THEN din END) avg_delta_in,
                 AVG(CASE WHEN dout IS NOT NULL THEN dout END) avg_delta_out,
                 AVG(CASE WHEN din IS NOT NULL THEN din*(COALESCE(payout_pct,65)/100.0)-COALESCE(dout,0) END) avg_period_gap,
                 SUM(CASE WHEN din IS NOT NULL AND din<=5 THEN 1 ELSE 0 END) low_moves,
                 MAX(CASE WHEN dout IS NOT NULL THEN dout ELSE 0 END) max_period_out
          FROM deltas GROUP BY codeid
        ), latest AS (SELECT * FROM deltas WHERE rn_desc=1),
        base AS (
          SELECT l.codeid,l.sede_nome,l.comune,l.stato,l.em,l.last_read_at,l.codice_modello,
                 COALESCE(NULLIF(l.modello,''),mc.nome_modello,'MODELLO NON CENSITO') modello,
                 l.cnt_in,l.cnt_out,
                 COALESCE(mc.ciclo,0) ciclo,
                 COALESCE(mc.payout,l.payout_pct,65) payout,
                 COALESCE(h.observations,0) observations,
                 COALESCE(h.avg_delta_in,0) avg_delta_in,
                 COALESCE(h.avg_delta_out,0) avg_delta_out,
                 COALESCE(h.avg_period_gap,0) avg_period_gap,
                 COALESCE(h.low_moves,0) low_moves,
                 COALESCE(h.max_period_out,0) max_period_out,
                 ROUND((l.cnt_in*(COALESCE(mc.payout,l.payout_pct,65)/100.0))-l.cnt_out,2) payout_gap,
                 ROUND(CASE WHEN COALESCE(mc.ciclo,0)>0 THEN l.cnt_in % mc.ciclo END,2) cycle_pos,
                 ROUND(CASE WHEN COALESCE(mc.ciclo,0)>0 THEN mc.ciclo-(l.cnt_in % mc.ciclo) END,2) cycle_remaining
          FROM latest l
          LEFT JOIN model_config mc ON mc.codice_modello=l.codice_modello
          LEFT JOIN hist h ON h.codeid=l.codeid
          WHERE UPPER(TRIM(COALESCE(l.stato,'')))='ATTIVA'
            AND UPPER(TRIM(COALESCE(l.em,'')))='E'
        ), model_stats AS (
          SELECT modello, AVG(avg_delta_in) model_avg_delta, AVG(payout_gap) model_avg_gap
          FROM base GROUP BY modello
        )
        SELECT b.*, COALESCE(ms.model_avg_delta,0) model_avg_delta,
               COALESCE(ms.model_avg_gap,0) model_avg_gap
        FROM base b LEFT JOIN model_stats ms ON ms.modello=b.modello
        """).fetchall()

    data=[]
    for rr in rows:
        r=dict(rr)
        cycle=float(r.get("ciclo") or 0)
        gap=float(r.get("payout_gap") or 0)
        position=float(r.get("cycle_pos") or 0)
        remain=float(r.get("cycle_remaining") or 0)
        avg_in=float(r.get("avg_delta_in") or 0)
        model_avg=float(r.get("model_avg_delta") or 0)
        observations=int(r.get("observations") or 0)
        low_moves=int(r.get("low_moves") or 0)

        # Indice gestionale/statistico, non probabilità matematica di vincita.
        score=50.0
        components=[]

        if cycle > 0:
            normalized_gap=max(-1.0,min(1.0,gap/max(cycle*0.20,1)))
            gap_points=normalized_gap*24
            score += gap_points
            components.append(("Scostamento payout", round(gap_points,1)))

            progress=max(0.0,min(1.0,position/cycle))
            progress_points=(progress-0.5)*18
            score += progress_points
            components.append(("Posizione nel ciclo", round(progress_points,1)))
        else:
            score -= 12
            components.append(("Ciclo non configurato", -12.0))

        if model_avg > 0:
            ratio=max(0.0,min(2.0,avg_in/model_avg))
            activity_points=(ratio-1.0)*10
            score += activity_points
            components.append(("Attività rispetto al modello", round(activity_points,1)))

        continuity=1-(low_moves/max(observations,1))
        continuity_points=(continuity-0.5)*10 if observations else -4
        score += continuity_points
        components.append(("Continuità di gioco", round(continuity_points,1)))

        freshness=read_status(r.get("last_read_at"))
        fresh_points={"green":6,"yellow":2,"orange":-4,"red":-10}.get(freshness["class"],-6)
        score += fresh_points
        components.append(("Aggiornamento dati", float(fresh_points)))

        if observations < 3:
            score -= 8
            components.append(("Storico insufficiente", -8.0))

        score=max(0,min(100,score))
        confidence=max(15,min(100,observations*8 + (20 if cycle>0 else 0)))

        reasons=[]; risks=[]
        if gap > max(cycle*0.03,300): reasons.append(f"OUT inferiore al teorico di circa € {gap:,.0f}")
        elif gap < -max(cycle*0.03,300): risks.append(f"OUT superiore al teorico di circa € {abs(gap):,.0f}")
        else: reasons.append("Scostamento dal payout teorico contenuto")
        if cycle>0:
            reasons.append(f"Ciclo stimato al {position/cycle*100:.0f}%")
        if model_avg>0 and avg_in>model_avg*1.15: reasons.append("Attività recente sopra la media del modello")
        if model_avg>0 and avg_in<model_avg*0.70: risks.append("Attività recente sotto la media del modello")
        if observations and low_moves/observations>0.35: risks.append("Molti intervalli con movimento minimo")
        if freshness["class"] in ("orange","red"): risks.append("Lettura non recente")
        if observations<3: risks.append("Pochi dati storici: indice meno affidabile")

        if score>=75: signal,label="green","Indice alto"
        elif score>=50: signal,label="orange","Indice medio"
        else: signal,label="red","Indice basso"

        estimated_in=None
        if gap>0 and float(r.get("payout") or 0)>0:
            estimated_in=gap/(float(r.get("payout"))/100.0)

        r.update(
            score=round(score,1), confidence=round(confidence), signal=signal,
            signal_label=label, read_status=freshness, reasons=reasons[:4], risks=risks[:4],
            components=sorted(components,key=lambda x:abs(x[1]),reverse=True),
            estimated_in=estimated_in,
            cycle_progress=(position/cycle*100 if cycle else None),
        )
        data.append(r)

    models=sorted({x["modello"] for x in data if x.get("modello")})
    locations=sorted({x["sede_nome"] for x in data if x.get("sede_nome")})
    municipalities=sorted({x["comune"] for x in data if x.get("comune")})

    def matches(x):
        if filters["model"] and x.get("modello") != filters["model"]: return False
        if filters["location"] and x.get("sede_nome") != filters["location"]: return False
        if filters["municipality"] and x.get("comune") != filters["municipality"]: return False
        if filters["band"] and x.get("signal") != filters["band"]: return False
        if filters["q"]:
            hay=" ".join(str(x.get(k) or "") for k in ("codeid","modello","sede_nome","comune")).lower()
            if filters["q"].lower() not in hay: return False
        return True

    filtered=[x for x in data if matches(x)]
    filtered.sort(key=lambda x:(x["score"],x["confidence"]),reverse=True)
    summary={
        "high":sum(1 for x in filtered if x["signal"]=="green"),
        "medium":sum(1 for x in filtered if x["signal"]=="orange"),
        "low":sum(1 for x in filtered if x["signal"]=="red"),
        "average":round(sum(x["score"] for x in filtered)/len(filtered),1) if filtered else 0,
    }
    return render_template("analysis.html",machines=filtered,summary=summary,filters=filters,
                           models=models,locations=locations,municipalities=municipalities)


def health_signal(score):
    if score >= 75:
        return 'green', 'Ottima'
    if score >= 50:
        return 'orange', 'Da monitorare'
    return 'red', 'Critica'


@app.route('/health')
def health():
    with db() as con:
        rows = con.execute("""
        WITH ordered AS (
          SELECT s.*,
                 LAG(cnt_in) OVER(PARTITION BY codeid ORDER BY snapshot_at,id) prev_in,
                 LAG(cnt_out) OVER(PARTITION BY codeid ORDER BY snapshot_at,id) prev_out,
                 ROW_NUMBER() OVER(PARTITION BY codeid ORDER BY snapshot_at DESC,id DESC) rn
          FROM snapshots s
        ), deltas AS (
          SELECT *, CASE WHEN prev_in IS NULL THEN NULL ELSE MAX(0,cnt_in-prev_in) END din,
                    CASE WHEN prev_out IS NULL THEN NULL ELSE MAX(0,cnt_out-prev_out) END dout
          FROM ordered
        ), hist AS (
          SELECT codeid, AVG(CASE WHEN din>0 THEN din END) avg_din,
                 SUM(CASE WHEN din<=5 THEN 1 ELSE 0 END) low_moves,
                 COUNT(din) observations, SUM(COALESCE(din,0)) total_played
          FROM deltas WHERE din IS NOT NULL GROUP BY codeid
        ), latest AS (SELECT * FROM deltas WHERE rn=1),
        model_perf AS (
          SELECT COALESCE(NULLIF(l.modello,''),mc.nome_modello,'MODELLO NON CENSITO') modello_ok,
                 AVG(COALESCE(h.avg_din,0)) model_avg
          FROM latest l LEFT JOIN hist h ON h.codeid=l.codeid
          LEFT JOIN model_config mc ON mc.codice_modello=l.codice_modello
          GROUP BY modello_ok
        )
        SELECT l.*, COALESCE(NULLIF(l.modello,''),mc.nome_modello,'MODELLO NON CENSITO') modello_ok,
               mc.ciclo, COALESCE(mc.payout,l.payout_pct,65) payout_cfg,
               COALESCE(h.avg_din,0) avg_din, COALESCE(h.low_moves,0) low_moves,
               COALESCE(h.observations,0) observations, COALESCE(h.total_played,0) total_played,
               COALESCE(mp.model_avg,0) model_avg
        FROM latest l
        LEFT JOIN hist h ON h.codeid=l.codeid
        LEFT JOIN model_config mc ON mc.codice_modello=l.codice_modello
        LEFT JOIN model_perf mp ON mp.modello_ok=COALESCE(NULLIF(l.modello,''),mc.nome_modello,'MODELLO NON CENSITO')
        WHERE UPPER(COALESCE(l.stato,'')) LIKE '%ATTIV%' AND UPPER(COALESCE(l.em,''))='E'
        """).fetchall()

    machines=[]
    for r in rows:
        d=dict(r)
        status=read_status(d.get('last_read_at'))
        freshness={'green':100,'yellow':75,'orange':45,'red':15}.get(status['class'],30)
        perf=50
        if d.get('model_avg',0)>0:
            perf=max(0,min(100,(d.get('avg_din',0)/d['model_avg'])*70))
        continuity=100
        if d.get('observations',0)>0:
            continuity=max(0,100-(d.get('low_moves',0)/d['observations']*100))
        config=100 if d.get('ciclo') and d.get('codice_modello') else 25
        score=round(freshness*.25+perf*.40+continuity*.25+config*.10,1)
        signal,label=health_signal(score)
        d.update({'score':score,'signal':signal,'health_label':label,'read_status':status,
                  'performance_pct':round((d.get('avg_din',0)/d['model_avg']*100) if d.get('model_avg',0) else 0,1),
                  'stopped_pct':round((d.get('low_moves',0)/d['observations']*100) if d.get('observations',0) else 0,1)})
        if score < 50:
            if d['stopped_pct']>=40:
                d['recommendation']='Verificare fermo, gettoniera, refill o disponibilità moneta'
            elif d['performance_pct'] and d['performance_pct']<60:
                d['recommendation']='Valutare spostamento o sostituzione: rendimento sotto media modello'
            else:
                d['recommendation']='Controllo operativo consigliato'
        elif score < 75:
            d['recommendation']='Monitorare nei prossimi sinottici'
        else:
            d['recommendation']='Nessuna criticità rilevante'
        machines.append(d)
    machines.sort(key=lambda x:x['score'])
    critical=sum(1 for m in machines if m['signal']=='red')
    monitor=sum(1 for m in machines if m['signal']=='orange')
    healthy=sum(1 for m in machines if m['signal']=='green')
    insights=[]
    if critical:
        insights.append(f'{critical} macchine richiedono un controllo prioritario.')
    stopped=sum(1 for m in machines if m['stopped_pct']>=40)
    if stopped:
        insights.append(f'{stopped} macchine mostrano contatori fermi o quasi fermi in almeno il 40% delle osservazioni.')
    under=sum(1 for m in machines if m['performance_pct'] and m['performance_pct']<60)
    if under:
        insights.append(f'{under} macchine rendono meno del 60% della media del proprio modello.')
    return render_template('health.html',machines=machines,critical=critical,monitor=monitor,healthy=healthy,insights=insights)

@app.route('/models', methods=['GET','POST'])
def models():
    if request.method == 'POST':
        code = str(request.form.get('codice_modello') or '').strip()
        name = str(request.form.get('nome_modello') or '').strip()
        ciclo = to_float(request.form.get('ciclo'))
        payout = to_float(request.form.get('payout')) or 65
        if code:
            with db() as con:
                con.execute('''INSERT OR REPLACE INTO model_config(codice_modello,nome_modello,ciclo,payout,updated_at) VALUES(?,?,?,?,CURRENT_TIMESTAMP)''', (code, name, ciclo, payout))
            flash('Modello/ciclo salvato.')
        return redirect(url_for('models'))
    with db() as con:
        rows = con.execute('''SELECT mc.*, COUNT(DISTINCT s.codeid) machines FROM model_config mc LEFT JOIN snapshots s ON s.codice_modello=mc.codice_modello GROUP BY mc.codice_modello ORDER BY mc.nome_modello''').fetchall()
        missing = con.execute('''SELECT s.codice_modello, COALESCE(NULLIF(s.modello,''),'Senza nome') modello, COUNT(DISTINCT s.codeid) machines FROM snapshots s LEFT JOIN model_config mc ON mc.codice_modello=s.codice_modello WHERE COALESCE(s.codice_modello,'')<>'' AND mc.codice_modello IS NULL GROUP BY s.codice_modello, s.modello ORDER BY machines DESC''').fetchall()
    return render_template('models.html', rows=rows, missing=missing)


@app.route('/sync-cicloslot', methods=['POST'])
def sync_cicloslot():
    n = sync_model_config_from_json()
    flash(f'Cicloslot sincronizzato: {n} modelli caricati/aggiornati.')
    return redirect(url_for('models'))


@app.route('/alarms')
def alarms():
    threshold = float(request.args.get('threshold', 5) or 5)
    only_stopped = request.args.get('stopped') == '1'
    with db() as con:
        missing_models = con.execute('''SELECT s.codice_modello, COALESCE(NULLIF(s.modello,''),'Senza nome') modello, COUNT(DISTINCT s.codeid) machines FROM snapshots s LEFT JOIN model_config mc ON mc.codice_modello=s.codice_modello WHERE COALESCE(s.codice_modello,'')<>'' AND mc.codice_modello IS NULL GROUP BY s.codice_modello, s.modello ORDER BY machines DESC''').fetchall()
        movement = con.execute('''
        WITH base AS (
          SELECT s.*, LAG(cnt_in) OVER(PARTITION BY codeid ORDER BY snapshot_at,id) prev_in, LAG(cnt_out) OVER(PARTITION BY codeid ORDER BY snapshot_at,id) prev_out,
                 ROW_NUMBER() OVER(PARTITION BY codeid ORDER BY snapshot_at DESC,id DESC) rn
          FROM snapshots s
        ), latest AS (
          SELECT *, MAX(0,cnt_in-prev_in) delta_in, MAX(0,cnt_out-prev_out) delta_out FROM base WHERE rn=1 AND prev_in IS NOT NULL
        ), hist AS (
          SELECT codeid, AVG(delta_in) avg_delta_in FROM (
            SELECT codeid, MAX(0,cnt_in-LAG(cnt_in) OVER(PARTITION BY codeid ORDER BY snapshot_at,id)) delta_in FROM snapshots
          ) WHERE delta_in IS NOT NULL AND delta_in>0 GROUP BY codeid
        )
        SELECT l.codeid,l.sede_nome,l.comune,l.codice_modello,COALESCE(NULLIF(l.modello,''),mc.nome_modello) modello, ROUND(l.delta_in,2) delta_in, ROUND(l.delta_out,2) delta_out, ROUND(COALESCE(h.avg_delta_in,0),2) avg_delta_in,
               CASE WHEN l.delta_out>=100 AND l.delta_in<=:threshold THEN 'Possibile fermo dopo forte pagamento / verificare hopper'
                    WHEN l.delta_in<=:threshold AND COALESCE(h.avg_delta_in,0)>:threshold THEN 'Contatori fermi o quasi fermi'
                    ELSE 'Da verificare' END alarm_text
        FROM latest l LEFT JOIN hist h ON h.codeid=l.codeid LEFT JOIN model_config mc ON mc.codice_modello=l.codice_modello
        WHERE (l.delta_in<=:threshold AND COALESCE(h.avg_delta_in,0)>:threshold) OR (l.delta_out>=100 AND l.delta_in<=:threshold)
        ORDER BY delta_out DESC, avg_delta_in DESC LIMIT 300
        ''', {'threshold': threshold}).fetchall()
    return render_template('alarms.html', missing_models=missing_models, movement=movement, threshold=threshold, only_stopped=only_stopped)


@app.route('/machine/<codeid>')
def machine(codeid):
    with db() as con:
        rows = con.execute('''SELECT s.*, COALESCE(NULLIF(s.modello,''),mc.nome_modello) modello_cfg, mc.ciclo, COALESCE(mc.payout,s.payout_pct,65) payout_cfg FROM snapshots s LEFT JOIN model_config mc ON mc.codice_modello=s.codice_modello WHERE codeid=? ORDER BY snapshot_at ASC, id ASC''', (codeid,)).fetchall()
        model_avg = con.execute('''
        WITH latest AS (SELECT *,ROW_NUMBER() OVER(PARTITION BY codeid ORDER BY snapshot_at DESC,id DESC) rn FROM snapshots),
        d AS (SELECT codeid,AVG(delta) avg_din FROM (SELECT codeid,MAX(0,cnt_in-LAG(cnt_in) OVER(PARTITION BY codeid ORDER BY snapshot_at,id)) delta FROM snapshots) WHERE delta IS NOT NULL AND delta>0 GROUP BY codeid)
        SELECT AVG(COALESCE(d.avg_din,0)) avg_model FROM latest l LEFT JOIN d ON d.codeid=l.codeid WHERE l.rn=1 AND l.codice_modello=(SELECT codice_modello FROM snapshots WHERE codeid=? ORDER BY snapshot_at DESC,id DESC LIMIT 1)
        ''',(codeid,)).fetchone()
    if not rows:
        return 'Macchina non trovata', 404
    timeline=[]; prev=None; deltas=[]
    for r in rows:
        if prev:
            din=max(0,(r['cnt_in'] or 0)-(prev['cnt_in'] or 0)); dout=max(0,(r['cnt_out'] or 0)-(prev['cnt_out'] or 0))
            deltas.append(din)
            if (r['sede_code'] or '') != (prev['sede_code'] or ''):
                timeline.append({'date':r['snapshot_at'],'title':'Spostamento locale','detail':f"Da {prev['sede_nome'] or '-'} a {r['sede_nome'] or '-'}"})
            if din <= 5:
                timeline.append({'date':r['snapshot_at'],'title':'Contatori fermi o quasi fermi','detail':f'Incremento IN {din:.2f} €, incremento OUT {dout:.2f} €'})
            if dout >= 100 and din <= 5:
                timeline.append({'date':r['snapshot_at'],'title':'Possibile fermo dopo pagamento','detail':f'OUT aumentato di {dout:.2f} € con IN quasi fermo: verificare hopper/refill'})
        prev=r
    latest=rows[-1]
    avg_din=sum(deltas)/len(deltas) if deltas else 0
    model_avg_value=(model_avg['avg_model'] if model_avg and model_avg['avg_model'] else 0)
    perf_pct=(avg_din/model_avg_value*100) if model_avg_value else 0
    return render_template('machine.html', codeid=codeid, rows=rows, latest=latest, timeline=list(reversed(timeline[-30:])), avg_din=avg_din, model_avg=model_avg_value, perf_pct=perf_pct, read_state=read_status(latest['last_read_at']))


@app.route('/api/machine/<codeid>')
def api_machine(codeid):
    with db() as con:
        rows = con.execute('''SELECT s.snapshot_at, s.cnt_in, s.cnt_out, COALESCE(mc.payout,s.payout_pct,65) payout FROM snapshots s LEFT JOIN model_config mc ON mc.codice_modello=s.codice_modello WHERE codeid=? ORDER BY snapshot_at ASC, id ASC''', (codeid,)).fetchall()
    data, prev = [], None
    for r in rows:
        inc_in = inc_out = 0
        if prev:
            inc_in = max(0, r['cnt_in'] - prev['cnt_in'])
            inc_out = max(0, r['cnt_out'] - prev['cnt_out'])
        p = (r['payout'] or 65) / 100.0
        data.append({'date': r['snapshot_at'], 'cnt_in': round(r['cnt_in'],2), 'cnt_out': round(r['cnt_out'],2), 'delta_in': round(inc_in,2), 'delta_out': round(inc_out,2), 'payout_real': round((r['cnt_out']/r['cnt_in']*100) if r['cnt_in'] else 0,2), 'out_to_realign_now': round((r['cnt_in']*p)-r['cnt_out'],2)})
        prev = r
    return jsonify(data)


@app.route('/stats')
def stats():
    with db() as con:
        totals = summary_totals(con)
        model_stats = con.execute('''
        WITH joined AS (
          SELECT s.*, COALESCE(NULLIF(s.modello,''), mc.nome_modello, 'MODELLO NON CENSITO') modello_ok
          FROM snapshots s LEFT JOIN model_config mc ON mc.codice_modello=s.codice_modello
        ), r AS (
          SELECT codeid, modello_ok modello, snapshot_at, cnt_in, cnt_out,
                 LAG(cnt_in) OVER(PARTITION BY codeid ORDER BY snapshot_at,id) prev_in,
                 LAG(cnt_out) OVER(PARTITION BY codeid ORDER BY snapshot_at,id) prev_out
          FROM joined
        ), d AS (
          SELECT modello, substr(snapshot_at,1,7) mese, MAX(0,cnt_in-prev_in) din, MAX(0,cnt_out-prev_out) dout FROM r WHERE prev_in IS NOT NULL
        ), m AS (
          SELECT modello, mese, SUM(din) mese_in, SUM(dout) mese_out FROM d GROUP BY modello,mese
        ), tot AS (
          SELECT modello, SUM(mese_in) total_in, AVG(mese_in) avg_month, MAX(mese_in) peak_in,
                 (SELECT mese FROM m mm WHERE mm.modello=m.modello ORDER BY mm.mese_in DESC LIMIT 1) peak_month,
                 (SELECT mese FROM m mm WHERE mm.modello=m.modello ORDER BY mm.mese DESC LIMIT 1) last_month,
                 (SELECT mese_in FROM m mm WHERE mm.modello=m.modello ORDER BY mm.mese DESC LIMIT 1) last_in
          FROM m GROUP BY modello
        )
        SELECT *, ROUND((last_in-avg_month),2) trend_vs_avg FROM tot WHERE total_in>0 ORDER BY total_in DESC LIMIT 100
        ''').fetchall()
        city_model = con.execute('''
        WITH joined AS (
          SELECT s.*, COALESCE(NULLIF(s.modello,''), mc.nome_modello, 'MODELLO NON CENSITO') modello_ok
          FROM snapshots s LEFT JOIN model_config mc ON mc.codice_modello=s.codice_modello
        ), r AS (
          SELECT codeid, modello_ok modello, comune, snapshot_at, cnt_in, LAG(cnt_in) OVER(PARTITION BY codeid ORDER BY snapshot_at,id) prev_in FROM joined WHERE comune<>''
        ), d AS (SELECT modello, comune, MAX(0,cnt_in-prev_in) din FROM r WHERE prev_in IS NOT NULL)
        SELECT modello, comune, ROUND(SUM(din),2) total_in FROM d GROUP BY modello, comune HAVING total_in>0 ORDER BY total_in DESC LIMIT 120
        ''').fetchall()
        top_city = con.execute('''
        WITH r AS (SELECT codeid, comune, snapshot_at, cnt_in, LAG(cnt_in) OVER(PARTITION BY codeid ORDER BY snapshot_at,id) prev_in FROM snapshots WHERE comune<>'')
        SELECT comune, ROUND(SUM(MAX(0,cnt_in-prev_in)),2) played FROM r WHERE prev_in IS NOT NULL GROUP BY comune ORDER BY played DESC LIMIT 10
        ''').fetchall()
    return render_template('stats.html', totals=totals, model_stats=model_stats, city_model=city_model, top_city=top_city)


@app.route('/api/model/<path:model>')
def api_model(model):
    with db() as con:
        rows = con.execute('''
        WITH joined AS (
          SELECT s.*, COALESCE(NULLIF(s.modello,''), mc.nome_modello, 'MODELLO NON CENSITO') modello_ok
          FROM snapshots s LEFT JOIN model_config mc ON mc.codice_modello=s.codice_modello
        ), r AS (
          SELECT codeid, snapshot_at, cnt_in, cnt_out,
                 LAG(cnt_in) OVER(PARTITION BY codeid ORDER BY snapshot_at,id) prev_in,
                 LAG(cnt_out) OVER(PARTITION BY codeid ORDER BY snapshot_at,id) prev_out
          FROM joined WHERE modello_ok=?
        ), d AS (SELECT substr(snapshot_at,1,7) mese, MAX(0,cnt_in-prev_in) din, MAX(0,cnt_out-prev_out) dout FROM r WHERE prev_in IS NOT NULL)
        SELECT mese, ROUND(SUM(din),2) in_mese, ROUND(SUM(dout),2) out_mese FROM d GROUP BY mese ORDER BY mese
        ''', (model,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/locations')
def locations():
    with db() as con:
        totals = summary_totals(con)
        locs = con.execute('''
        WITH latest AS (SELECT *, ROW_NUMBER() OVER(PARTITION BY codeid ORDER BY snapshot_at DESC,id DESC) rn FROM snapshots),
        r AS (SELECT codeid, sede_code, snapshot_at, cnt_in, LAG(cnt_in) OVER(PARTITION BY codeid ORDER BY snapshot_at,id) prev_in FROM snapshots),
        d AS (SELECT sede_code, SUM(MAX(0,cnt_in-prev_in)) played FROM r WHERE prev_in IS NOT NULL GROUP BY sede_code)
        SELECT l.sede_code, l.sede_nome, l.comune, COUNT(*) active_machines, ROUND(COALESCE(d.played,0),2) played FROM latest l LEFT JOIN d ON d.sede_code=l.sede_code WHERE l.rn=1 GROUP BY l.sede_code,l.sede_nome,l.comune ORDER BY played DESC
        ''').fetchall()
    return render_template('locations.html', totals=totals, locations=locs)


@app.route('/location/<sede_code>')
def location_detail(sede_code):
    with db() as con:
        loc = con.execute('SELECT sede_code,sede_nome,indirizzo,comune,provincia FROM snapshots WHERE sede_code=? ORDER BY snapshot_at DESC LIMIT 1', (sede_code,)).fetchone()
        machines = con.execute('''
        WITH latest AS (SELECT *, ROW_NUMBER() OVER(PARTITION BY codeid ORDER BY snapshot_at DESC,id DESC) rn FROM snapshots),
        r AS (SELECT codeid, snapshot_at, cnt_in, LAG(cnt_in) OVER(PARTITION BY codeid ORDER BY snapshot_at,id) prev_in FROM snapshots),
        d AS (SELECT codeid, SUM(MAX(0,cnt_in-prev_in)) played FROM r WHERE prev_in IS NOT NULL GROUP BY codeid),
        model_avg AS (SELECT modello, AVG(played) avg_model_played FROM (SELECT l.codeid, l.modello, COALESCE(d.played,0) played FROM latest l LEFT JOIN d ON d.codeid=l.codeid WHERE l.rn=1) GROUP BY modello)
        SELECT l.codeid,l.codice_modello,l.modello,l.stato,l.cnt_in,l.cnt_out,COALESCE(d.played,0) played, ma.avg_model_played, ROUND(COALESCE(d.played,0)-COALESCE(ma.avg_model_played,0),2) diff_vs_model FROM latest l LEFT JOIN d ON d.codeid=l.codeid LEFT JOIN model_avg ma ON ma.modello=l.modello WHERE l.rn=1 AND l.sede_code=? ORDER BY diff_vs_model ASC
        ''', (sede_code,)).fetchall()
        historic = con.execute('SELECT DISTINCT codeid, modello, MIN(snapshot_at) first_seen, MAX(snapshot_at) last_seen FROM snapshots WHERE sede_code=? GROUP BY codeid,modello ORDER BY last_seen DESC', (sede_code,)).fetchall()
    return render_template('location_detail.html', loc=loc, machines=machines, historic=historic)


@app.route('/reset', methods=['POST'])
def reset():
    with db() as con:
        con.execute('DELETE FROM snapshots')
        con.execute('DELETE FROM imported_files')
        con.execute('DELETE FROM import_logs')
    flash('Archivio sinottici svuotato. I cicli/modelli restano salvati.')
    return redirect(url_for('index'))


init_db()
sync_model_config_from_json()

if __name__ == '__main__':
    app.run(debug=True)
